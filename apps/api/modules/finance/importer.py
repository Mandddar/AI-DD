"""
Financial data importer — parses Excel/TSV, detects chart of accounts,
maps accounts to standardized categories, handles German number formats.
"""
import csv
import io
import re
import logging
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from .models import (
    FinancialDataset, FinancialLineItem, AccountMapping,
    ChartOfAccounts, DatasetStatus,
)

logger = logging.getLogger(__name__)

# Month name patterns (German + English)
MONTH_PATTERNS = {
    "jan": 1, "januar": 1, "january": 1,
    "feb": 2, "februar": 2, "february": 2,
    "mär": 3, "mar": 3, "märz": 3, "march": 3,
    "apr": 4, "april": 4,
    "mai": 5, "may": 5,
    "jun": 6, "juni": 6, "june": 6,
    "jul": 7, "juli": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "okt": 10, "oct": 10, "oktober": 10, "october": 10,
    "nov": 11, "november": 11,
    "dez": 12, "dec": 12, "dezember": 12, "december": 12,
}


@dataclass
class ImportStructure:
    header_row: int = 0
    account_code_col: int = 0
    account_name_col: int = 1
    period_columns: dict = field(default_factory=dict)  # col_index -> date
    is_german_format: bool = False


def parse_german_number(value: str) -> Decimal:
    """Parse German number format: 1.234.567,89 → 1234567.89"""
    v = str(value).strip()
    if not v or v == "-":
        return Decimal("0")
    # German format: dots as thousands separator, comma as decimal
    if "," in v and "." in v:
        v = v.replace(".", "").replace(",", ".")
    elif "," in v:
        v = v.replace(",", ".")
    try:
        return Decimal(v)
    except InvalidOperation:
        return Decimal("0")


def parse_number(value, is_german: bool = False) -> Decimal:
    """Parse a number from a cell value."""
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    s = str(value).strip()
    if not s or s == "-":
        return Decimal("0")
    if is_german:
        return parse_german_number(s)
    try:
        return Decimal(s.replace(",", ""))
    except InvalidOperation:
        return Decimal("0")


def parse_period(header) -> date | None:
    """Try to parse a column header as a period date."""
    if header is None:
        return None
    s = str(header).strip().lower()

    # Pattern: "01/2024", "1/2024", "01.2024"
    m = re.match(r"(\d{1,2})[./](\d{4})", s)
    if m:
        return date(int(m.group(2)), int(m.group(1)), 1)

    # Pattern: "2024-01"
    m = re.match(r"(\d{4})-(\d{1,2})", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), 1)

    # Pattern: "Jan 2024", "Januar 2024", "January 2024"
    for name, month_num in MONTH_PATTERNS.items():
        if s.startswith(name):
            year_match = re.search(r"(\d{4})", s)
            if year_match:
                return date(int(year_match.group(1)), month_num, 1)

    # Pattern: "Jan-24", "Mar-23"
    m = re.match(r"([a-zäö]+)[- ]?(\d{2})$", s)
    if m:
        month_name = m.group(1)
        year_short = int(m.group(2))
        year = 2000 + year_short if year_short < 100 else year_short
        for name, month_num in MONTH_PATTERNS.items():
            if month_name.startswith(name):
                return date(year, month_num, 1)

    return None


def detect_structure(rows: list[list]) -> ImportStructure:
    """Analyze first rows to detect spreadsheet structure."""
    structure = ImportStructure()

    # Find header row: row with the most string cells that look like headers
    best_row = 0
    best_score = 0
    for i, row in enumerate(rows[:20]):
        score = 0
        for cell in row:
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if any(kw in s for kw in ["konto", "account", "bezeichnung", "name", "nr"]):
                score += 3
            period = parse_period(cell)
            if period:
                score += 2
            if isinstance(cell, str) and len(cell) > 2:
                score += 1
        if score > best_score:
            best_score = score
            best_row = i

    structure.header_row = best_row
    header = rows[best_row] if best_row < len(rows) else []

    # Find account code column (4-digit numbers in data rows below header)
    for col_idx in range(min(5, len(header))):
        digit_count = 0
        for row in rows[best_row + 1: best_row + 10]:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip()
                if re.match(r"^\d{4,5}$", val):
                    digit_count += 1
        if digit_count >= 3:
            structure.account_code_col = col_idx
            structure.account_name_col = col_idx + 1
            break

    # Find period columns
    for col_idx, cell in enumerate(header):
        period = parse_period(cell)
        if period:
            structure.period_columns[col_idx] = period

    # Detect German number format from data cells
    for row in rows[best_row + 1: best_row + 10]:
        for col_idx in structure.period_columns:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx])
                # German format: comma followed by exactly 2 digits at end
                if re.search(r",\d{2}$", val) and "." in val:
                    structure.is_german_format = True
                    break
        if structure.is_german_format:
            break

    return structure


def detect_chart_of_accounts(account_codes: list[str], amounts: dict[str, Decimal]) -> ChartOfAccounts:
    """Detect SKR03 vs SKR04 based on revenue account ranges."""
    has_4xxx_revenue = False
    has_8xxx_revenue = False

    for code in account_codes:
        if not code or not code.isdigit():
            continue
        num = int(code)
        amt = amounts.get(code, Decimal("0"))
        # Revenue accounts typically have positive amounts
        if 4000 <= num <= 4999 and amt > 0:
            has_4xxx_revenue = True
        if 8000 <= num <= 8999 and amt > 0:
            has_8xxx_revenue = True

    if has_8xxx_revenue and not has_4xxx_revenue:
        return ChartOfAccounts.skr04
    if has_4xxx_revenue and not has_8xxx_revenue:
        return ChartOfAccounts.skr03
    return ChartOfAccounts.custom


def read_excel(file_bytes: bytes) -> list[list]:
    """Read Excel file into a list of rows."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def read_tsv(file_bytes: bytes) -> list[list]:
    """Read TSV file into a list of rows."""
    text_content = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text_content), delimiter="\t")
    return [list(row) for row in reader]


async def map_accounts(
    line_items: list[FinancialLineItem],
    chart_type: ChartOfAccounts,
    db: AsyncSession,
) -> None:
    """Map account codes to standardized categories using the account_mappings table."""
    result = await db.execute(
        select(AccountMapping).where(AccountMapping.chart_type == chart_type)
    )
    mappings = list(result.scalars().all())

    for item in line_items:
        if not item.account_code or not item.account_code.isdigit():
            continue
        code_num = int(item.account_code)
        for m in mappings:
            try:
                start = int(m.account_code_start)
                end = int(m.account_code_end)
                if start <= code_num <= end:
                    item.standardized_category = m.standardized_category
                    break
            except ValueError:
                continue


async def process_financial_import(
    dataset_id: UUID,
    file_bytes: bytes,
    filename: str,
    project_id: UUID,
) -> None:
    """Background task: parse file, detect structure, map accounts, insert line items."""
    from core.database import AsyncSessionLocal

    async with AsyncSessionLocal() as db:
        dataset = await db.get(FinancialDataset, dataset_id)
        if not dataset:
            return

        try:
            # Parse file
            ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
            if ext in ("xlsx", "xls"):
                rows = read_excel(file_bytes)
            elif ext in ("tsv", "txt"):
                rows = read_tsv(file_bytes)
            else:
                raise ValueError(f"Unsupported file format: {ext}")

            if len(rows) < 2:
                raise ValueError("File has too few rows to contain financial data")

            # Detect structure
            structure = detect_structure(rows)
            if not structure.period_columns:
                raise ValueError("Could not detect period columns. Ensure headers contain month/year labels.")

            # Extract line items
            items = []
            account_amounts: dict[str, Decimal] = {}

            for row_idx in range(structure.header_row + 1, len(rows)):
                row = rows[row_idx]
                if structure.account_code_col >= len(row):
                    continue

                code = str(row[structure.account_code_col] or "").strip()
                if not code or not any(c.isdigit() for c in code):
                    continue

                name = ""
                if structure.account_name_col < len(row):
                    name = str(row[structure.account_name_col] or "").strip()

                for col_idx, period_date in structure.period_columns.items():
                    if col_idx >= len(row):
                        continue
                    amount = parse_number(row[col_idx], structure.is_german_format)
                    if amount == 0:
                        continue

                    items.append(FinancialLineItem(
                        project_id=project_id,
                        dataset_id=dataset_id,
                        account_code=code,
                        account_name=name,
                        period=period_date,
                        amount=amount,
                    ))
                    # Track amounts for chart detection
                    account_amounts[code] = account_amounts.get(code, Decimal("0")) + amount

            if not items:
                raise ValueError("No financial data rows found in file")

            # Detect chart of accounts
            codes = list(set(item.account_code for item in items))
            chart = detect_chart_of_accounts(codes, account_amounts)
            dataset.chart_of_accounts = chart

            # Map accounts to categories
            await map_accounts(items, chart, db)

            # Insert line items (upsert: newer data wins)
            for item in items:
                await db.execute(
                    text("""
                        INSERT INTO financial_line_items
                            (id, project_id, dataset_id, account_code, account_name,
                             standardized_category, period, amount, currency, created_at)
                        VALUES
                            (gen_random_uuid(), :project_id, :dataset_id, :account_code, :account_name,
                             :standardized_category, :period, :amount, :currency, now())
                        ON CONFLICT (project_id, account_code, period)
                        DO UPDATE SET
                            amount = EXCLUDED.amount,
                            dataset_id = EXCLUDED.dataset_id,
                            account_name = EXCLUDED.account_name,
                            standardized_category = EXCLUDED.standardized_category
                    """).bindparams(
                        project_id=project_id,
                        dataset_id=dataset_id,
                        account_code=item.account_code,
                        account_name=item.account_name,
                        standardized_category=item.standardized_category,
                        period=item.period,
                        amount=float(item.amount),
                        currency=item.currency,
                    )
                )

            # Update dataset metadata
            periods = sorted(set(item.period for item in items))
            dataset.period_start = periods[0]
            dataset.period_end = periods[-1]
            dataset.row_count = len(items)
            dataset.status = DatasetStatus.completed
            await db.commit()

        except Exception as e:
            logger.exception("Financial import failed for dataset %s: %s", dataset_id, e)
            dataset.status = DatasetStatus.failed
            dataset.error_message = str(e)[:500]
            await db.commit()
